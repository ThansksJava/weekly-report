"""Excel 导出 —— 使用 openpyxl 将周报渲染为带边框样式的工作簿。"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import Report

THIN = Side(style="thin", color="333333")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
WRAP = Alignment(vertical="top", wrap_text=True)
HEADER_ALIGN = Alignment(vertical="center", horizontal="center", wrap_text=True)


def report_to_xlsx(report: Report) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Report"

    row = 1
    ws.cell(row=row, column=1, value=report.title).font = Font(size=14, bold=True)
    row += 2
    ws.cell(row=row, column=1, value=report.greeting).font = Font(size=11)
    row += 1
    ws.cell(row=row, column=1, value=report.subtitle).font = Font(size=12, bold=True)
    row += 2

    max_cols = 1
    for section in report.sections:
        if section.name:
            ws.cell(row=row, column=1, value=section.name).font = Font(size=12, bold=True)
            row += 1

        col_offset = 0
        if section.show_index:
            cell = ws.cell(row=row, column=1, value="No.")
            cell.font = Font(bold=True)
            cell.border = BORDER
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            ws.column_dimensions["A"].width = 6
            col_offset = 1

        for i, col in enumerate(section.columns):
            cell = ws.cell(row=row, column=col_offset + i + 1, value=col.label)
            cell.font = Font(bold=True)
            cell.border = BORDER
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            letter = get_column_letter(col_offset + i + 1)
            # 像素宽度近似转 Excel 字符宽度
            ws.column_dimensions[letter].width = max(10, col.width // 8)
        row += 1

        for idx, data in enumerate(section.rows, start=1):
            if section.show_index:
                cell = ws.cell(row=row, column=1, value=idx)
                cell.border = BORDER
                cell.alignment = WRAP
            for i, col in enumerate(section.columns):
                cell = ws.cell(row=row, column=col_offset + i + 1, value=data.get(col.key, ""))
                cell.border = BORDER
                cell.alignment = WRAP
            row += 1

        max_cols = max(max_cols, col_offset + len(section.columns))
        row += 2  # 区块间空行

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
