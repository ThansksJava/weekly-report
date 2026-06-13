"""历史周报 Excel 导入 —— 解析本系统导出的 .xlsx,同时尽力兼容团队原有格式。

启发式规则:
- 第一个非空行视为标题;表格出现之前的单独文本行依次视为问候语/副标题
- 某行有 >=3 个非空单元格视为表头行;其上方最近的单独文本行作为区块名
- 表头行之后的行为数据行,直到完全空行结束(纯表头延续行——所有数据列为空——会被跳过)
- 表头首列为 "No." 时启用序号列
"""
from __future__ import annotations

import datetime as dt
import io
import re

from openpyxl import load_workbook

from .models import Report, Section, Column, new_id


def _cell_text(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def _find_dates(text: str) -> tuple[str, str] | None:
    """从标题中提取形如 2026.05.11- 2026.05.15 的日期范围。"""
    m = re.findall(r"(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})", text)
    if len(m) >= 2:
        def iso(t):
            return f"{t[0]}-{int(t[1]):02d}-{int(t[2]):02d}"
        return iso(m[0]), iso(m[1])
    return None


def parse_xlsx(data: bytes, user_id: str) -> Report:
    wb = load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active
    grid = [[_cell_text(c) for c in row] for row in ws.iter_rows(values_only=True)]

    title, greeting, subtitle = "", "", ""
    sections: list[Section] = []
    pending_name = ""  # 表格上方最近的单独文本行
    i = 0
    n = len(grid)

    while i < n:
        row = grid[i]
        filled = [c for c in row if c]
        if not filled:
            i += 1
            continue

        if len(filled) >= 3:  # 表头行 → 开始一个区块
            header = row
            show_index = header[0].strip().lower() in ("no.", "no", "序号", "#")
            start_col = 1 if show_index else 0
            # 表头列:从 start_col 起,到最后一个非空表头
            last = max(idx for idx, c in enumerate(header) if c)
            labels = [header[c] or f"列{c}" for c in range(start_col, last + 1)]
            columns = [Column(key=f"c{ci}", label=lab, width=160) for ci, lab in enumerate(labels)]

            rows: list[dict[str, str]] = []
            i += 1
            while i < n:
                drow = grid[i]
                if not any(c for c in drow):
                    break
                values = {f"c{ci}": (drow[start_col + ci] if start_col + ci < len(drow) else "")
                          for ci in range(len(columns))}
                # 跳过多行表头的延续行(仅个别列有文字、与表头同样式难判断;
                # 简化处理:首个数据列与序号列同时为空且非空格数 <=1 视为延续)
                non_empty = sum(1 for v in values.values() if v)
                idx_cell = drow[0] if show_index else ""
                if non_empty <= 1 and not idx_cell and rows == [] and non_empty > 0:
                    i += 1
                    continue
                rows.append(values)
                i += 1
            if not rows:
                rows = [{}]
            sections.append(Section(
                id=new_id(), name=pending_name, columns=columns, rows=rows, show_index=show_index,
            ))
            pending_name = ""
            continue

        # 单独文本行:按顺序识别标题/问候语/副标题,之后视为下一个区块名
        text = " ".join(filled)
        if not title:
            title = text
        elif not greeting:
            greeting = text
        elif not subtitle:
            subtitle = text
        else:
            pending_name = text
        i += 1

    dates = _find_dates(title) or _find_dates(subtitle)
    if dates:
        week_start, week_end = dates
    else:
        today = dt.date.today()
        monday = today - dt.timedelta(days=today.weekday())
        week_start, week_end = monday.isoformat(), (monday + dt.timedelta(days=6)).isoformat()

    return Report(
        id=new_id(), user_id=user_id,
        title=title or "Imported Weekly Report",
        greeting=greeting, subtitle=subtitle,
        week_start=week_start, week_end=week_end,
        sections=sections or [Section(id=new_id(), name="导入内容", columns=[], rows=[])],
        updated_at=dt.datetime.now().isoformat(timespec="seconds"),
    )
